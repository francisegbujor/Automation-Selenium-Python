import openpyxl

class HomePageData:

    test_HomePage_data = [{"name" : "Rahul", "email" : "shetty", "gender" : "Male"}, {"name" : "Demo", "email" : "hello@gmail", "gender" : "Female"}]

    @staticmethod
    def getExcelData(test_case_name):
        Dict = {}
        book = openpyxl.load_workbook("C:\\Users\\Francis\\Documents\\PythonDemo.xlsx")
        sheet = book.active
        for i in range(1, sheet.max_row + 1):  # to get rows
            if sheet.cell(row=i, column=1).value == test_case_name:

                for j in range(2, sheet.max_column + 1):  # to get columns
                    Dict[sheet.cell(row=1, column=j).value] = sheet.cell(row=i, column=j).value
        return[Dict]